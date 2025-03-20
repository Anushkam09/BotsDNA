from RPA.Browser.Selenium import Selenium
import os
from dotenv import load_dotenv
from openpyxl import load_workbook
from RPA.Email.ImapSmtp import ImapSmtp

class Mail:
    def __init__(self):
        self.mail = ImapSmtp()
        self.recipient = os.getenv("GMAIL_ACCOUNT")
        self.manager_email = os.getenv("GMAIL_ACCOUNT")
        self.gmail_account = os.getenv("GMAIL_ACCOUNT")
        self.gmail_password = os.getenv("GMAIL_PASSWORD")

        self.mail.authorize(
            account=self.gmail_account,
            password=self.gmail_password,
            smtp_server="smtp.gmail.com",
            smtp_port=587,
        )
    
    def send_mail(self, data ,employee_mail= None, manager_email=None):
        formatted_data = self.format_data(data)
        if employee_mail is None or manager_email is None:
            recipient = self.recipient
            manager_email = self.manager_email
        else:
            recipient = employee_mail
        self.mail.send_message(
            recipients=recipient,
            sender= self.gmail_account,
            cc=manager_email,
            subject= f"{data['employee_name']}, Happy Anniversary",
            body=formatted_data,
            html=True
        )
        print("mail_sent to ", data['employee_name'])
        return
    
    def format_data(self, data):
        data_to_send = """
        <!DOCTYPE html>
<html lang="en">
  <head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>Certificate</title>
    <link rel="preconnect" href="https://fonts.googleapis.com" />
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin />
    <link
      href="https://fonts.googleapis.com/css2?family=Montserrat:ital,wght@0,100..900;1,100..900&display=swap"
      rel="stylesheet"
    />
    <link rel="preconnect" href="https://fonts.googleapis.com" />
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin />
    <link
      href="https://fonts.googleapis.com/css2?family=DM+Serif+Text:ital@0;1&family=Montserrat:ital,wght@0,100..900;1,100..900&display=swap"
      rel="stylesheet"
    />
    <style>
      .compname {
        color: red;
        letter-spacing: 2px;
        text-align: center;
        padding: 10px;
      }
      .line {
        border: none;
        height: 2.1px;
        background-color: red;
        margin: 10px;
      }
      .anniversary {
        text-align: center;
        font-weight: 900;
        margin: 40px;
        font-size: 60px;
        padding: 10px;
      }
      .awardline {
        text-align: center;
        padding: 10px;
        margin: 50px;
        font-size: 20px;
      }
      .employee {
        color: red;
        letter-spacing: 1px;
        font-weight: bold;
        padding: 10px;
      }
      .employee {
        text-align: center;
        padding: 10px;
        margin: 10px;
      }
      .content {
        text-align: center;
        letter-spacing: 0.7px;
        word-spacing: 1px;
        padding: 10px;
        margin: 51.8px;
      }
      #montserrat-fonter {
        font-family: "Montserrat", serif;
        font-optical-sizing: auto;
        font-weight: 600;
        font-style: normal;
      }
      #dm-serif-text-regular {
        font-family: "DM Serif Text", serif;
        font-weight: 400;
        font-style: normal;
      }
    </style>
  </head> """ + f"""
  <body>
    <h2 id="montserrat-fonter" class="compname">BotsDNA</h2>
    <hr class="line" />
    <b><h1 id="dm-serif-text-regular" class="anniversary">Happy work Anniversary!</h1></b>
    <h3  class="awardline">This certificate is awarderd to</h3>
    <h2 class="employee">{data['employee_name']}</h2>
    <h3 class="content">
      for [his/her] outstanding service,tireless effort, constant support for
      BotsDNA and its projects for the last {2024 - int(data["date_of_join"].strip()[-4:])} years.
    </h3>
  </body>
</html>
        """
        return data_to_send


class ExcelHandling:
    def __init__(self, file_path):
        self.file_path = file_path
    
    def get_email_id(self, employee_id):
        wb = load_workbook(self.file_path)
        ws = wb.active

        for row in ws.iter_rows(min_row= 2, max_col=ws.max_column, max_row=ws.max_row, values_only=True):
            if row[0] == employee_id:
                return row[2]
        return None


class WorkAnniversarySite:
    def __init__(self,url, download_directory):
        self.url = url
        self.download_directory = download_directory
        self.browser = Selenium()
        self.browser.set_download_directory(self.download_directory)
        self.browser.open_chrome_browser(self.url)
    
    def download_excel(self):
        self.browser.wait_until_element_is_visible("//a[contains(@href, 'WorkAnniversary.xlsx')]")
        self.browser.click_element("//a[contains(@href, 'WorkAnniversary.xlsx')]")
        self.waiting(f"{self.download_directory}/WorkAnniversary.xlsx")
    
    def extract_details_from_page(self):
        rows = self.browser.find_elements("//html/body/center/table/tbody/tr")[1:]
        details = []
        for index, row in enumerate(rows):
            employee = {}
            employee["employee_id"] = self.browser.get_text(f"//html/body/center/table/tbody/tr[{index+2}]/td[1]")
            employee["employee_name"] = self.browser.get_text(f"//html/body/center/table/tbody/tr[{index+2}]/td[2]")
            employee["manager_id"] = self.browser.get_text(f"//html/body/center/table/tbody/tr[{index+2}]/td[3]")
            employee["date_of_join"] = self.browser.get_text(f"//html/body/center/table/tbody/tr[{index+2}]/td[4]")
            details.append(employee)
            
        
        return details
    
    def go_to_submit_page(self):
        self.browser.go_to(f"{self.url}/YearsOfExp.html")

    def submit_details(self, surname, name, years, manager_id):
        self.browser.wait_until_element_is_visible("id:surname")
        self.browser.input_text("id:surname", surname)
        self.browser.input_text("id:name", name)
        self.browser.input_text("id:years_of_experience", years)
        self.browser.input_text("id:manager_id", manager_id)
        self.browser.click_element('//*[@id="submission-form"]/input[5]')

    def waiting(self, file_path):
        while not os.path.exists(file_path):
            continue
        return
    
class Process:
    def __init__(self, url, download_directory):
        self.url = url
        self.download_directory = download_directory
        
    
    def work_anniversary(self):
        website = WorkAnniversarySite(self.url, self.download_directory)
        website.download_excel()
        details = website.extract_details_from_page()
        excel = ExcelHandling(f"{self.download_directory}/WorkAnniversary.xlsx")
        mail_work = Mail()
        for data in details:
            employee_email_id = excel.get_email_id(data["employee_id"])
            manager_email_id = excel.get_email_id(data["manager_id"])
            mail_work.send_mail(data, employee_email_id, manager_email_id)

        website.go_to_submit_page()

        for data in details:
            surname = data["employee_name"].strip().split()[-1]
            name = data["employee_name"].strip().split()[:-1]
            years_of_exp = 2024 - int(data["date_of_join"].strip()[-4:])
            manager_id = data["manager_id"]
            website.submit_details(surname,name,years_of_exp,manager_id)


if __name__ == "__main__":
    URL = "https://botsdna.com/WorkAnniversary/"
    DOWNLOAD_DIRECTORY = "./output"

    load_dotenv()

    process = Process(URL, DOWNLOAD_DIRECTORY)
    process.work_anniversary()
