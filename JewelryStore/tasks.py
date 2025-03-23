from RPA.Browser.Selenium import Selenium
from openpyxl import load_workbook
import os

class JewelryStoreSite:
    def __init__(self, url, download_directory):
        self.url = url
        self.download_directory = download_directory
        self.browser = Selenium()
        self.browser.set_download_directory(download_directory)
        self.browser.open_chrome_browser(url= self.url)
    
    def download_excel_file(self):
        self.browser.wait_until_element_is_visible("//a[contains(@href, 'CurrentMonthNewJewelry.xlsx')]")
        self.browser.click_element("//a[contains(@href, 'CurrentMonthNewJewelry.xlsx')]")
        self.waiting(f"{self.download_directory}/CurrentMonthNewJewelry.xlsx")

    def fill_form(self, data, category):
        self.browser.wait_until_element_is_visible("id:ddlCategory")
        for jewelry_type, metal_type in data.items():
            self.browser.click_element("id:ddlCategory")
            self.browser.select_from_list_by_value("id:ddlCategory", category)
            checkboxes = self.browser.find_elements("//td/input[@type='checkbox']")
            checkboxes_text = [element.text for element in self.browser.find_elements("//td")][-18:-2]
            self.browser.select_from_list_by_value("id:ddlJewelry", jewelry_type)
            for index, text in enumerate(checkboxes_text):
                if text in metal_type:
                    try:
                        self.browser.click_element(checkboxes[index])
                    except Exception as e:
                        print("Error", e, text)
            self.browser.click_element('//*[@id="courts"]/tbody/tr[4]/td/input')
            self.browser.handle_alert()
            self.browser.go_to(url= self.url)
    
    def waiting(self, file_path):
        while not os.path.exists(file_path):
            continue
        return

class ExcelHandling:
    def __init__(self, file_name):
        self.file_name = file_name
        self.wb = load_workbook(filename= self.file_name)
    
    def get_sheets(self):
        return self.wb.sheetnames
    
    def get_data_from_sheet(self, sheet):
        data_dict = {}
        active_sheet = self.wb[sheet]
        for row in active_sheet.iter_rows(min_row=2, max_row=active_sheet.max_row, max_col=2, values_only=True):
            jewelry_type, metal_type = row
            if jewelry_type in data_dict:
                data_dict[jewelry_type].append(metal_type)
            else:
                data_dict[jewelry_type] = [metal_type]
        return data_dict
        
class Process:
    def __init__(self,url, download_directory, file_name):
        self.url = url
        self.download_directory = download_directory
        self.file_name = file_name
    
    def JewelryStore(self):
        website = JewelryStoreSite(self.url, self.download_directory)
        website.download_excel_file()

        Excel = ExcelHandling(self.file_name)
        sheets = Excel.get_sheets()
        for sheet in sheets:
            if "_" in sheet:
                category = sheet.replace("_", "'s ")
            else:
                category = sheet
            data = Excel.get_data_from_sheet(sheet)
            website.fill_form(data, category)
            
        
if __name__ == "__main__":
    URL = "https://botsdna.com/jewelry/"
    DOWNLOAD_DIRECTORY = "./JewelryStore/output"
    FILE_NAME = "./JewelryStore/output/CurrentMonthNewJewelry.xlsx"
    process = Process(URL, DOWNLOAD_DIRECTORY, FILE_NAME)
    process.JewelryStore()

    
